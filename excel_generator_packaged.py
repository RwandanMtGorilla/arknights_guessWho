# excel_generator_packaged.py
# 适用于PyInstaller打包的Excel生成器版本 - 修改版
# 生成文件保存在exe文件同一目录

import json
import os
import sys
import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill
from PIL import Image as PILImage
import tempfile

# 尝试导入图片功能
try:
    from openpyxl.drawing.image import Image
    HAS_IMAGE_SUPPORT = True
except ImportError:
    HAS_IMAGE_SUPPORT = False
    print("警告: 无法导入图片功能，将使用文字版本")

def get_resource_path(relative_path):
    """
    获取资源文件的绝对路径，支持PyInstaller打包后的路径
    """
    try:
        # PyInstaller创建临时文件夹，将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
    except Exception:
        # 如果不是打包后的程序，使用脚本所在目录
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

def get_exe_directory():
    """
    获取exe文件所在的目录路径
    """
    if getattr(sys, 'frozen', False):
        # 如果是打包后的exe文件
        exe_dir = os.path.dirname(sys.executable)
    else:
        # 如果是脚本运行模式
        exe_dir = os.path.dirname(os.path.abspath(__file__))
    
    return exe_dir

def check_and_create_output_dir():
    """
    检查并创建输出目录（exe文件同一目录）
    """
    # 获取exe文件所在目录作为输出目录
    output_dir = get_exe_directory()
    
    # 检查目录写入权限
    try:
        test_file = os.path.join(output_dir, 'test_write_permission.tmp')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        print(f"✅ 输出目录: {output_dir}")
    except Exception as e:
        print(f"⚠️ 当前目录写入权限检查失败: {e}")
        print(f"🔄 仍将尝试在此目录生成文件: {output_dir}")
    
    return output_dir

def get_avatar_filename(operator_name, avatars_folder):
    """
    根据干员姓名获取正确的头像文件名
    格式: {姓名}_头像_{姓名}.png
    """
    # 基本格式
    filename = f"{operator_name}_头像_{operator_name}.png"
    file_path = os.path.join(avatars_folder, filename)
    
    if os.path.exists(file_path):
        return filename, file_path
    
    # 如果基本格式不存在，尝试其他可能的格式
    possible_formats = [
        f"头像_{operator_name}.png",  # 原始格式
        f"{operator_name}.png",       # 简单格式
        f"{operator_name}_头像.png",  # 变体格式
    ]
    
    for format_name in possible_formats:
        test_path = os.path.join(avatars_folder, format_name)
        if os.path.exists(test_path):
            return format_name, test_path
    
    return None, None

def get_valid_operators(json_file, avatars_folder):
    """
    获取所有有效的干员数据（有头像文件的）
    """
    if not os.path.exists(json_file):
        print(f"❌ 错误: 找不到文件 {json_file}")
        return []
    
    if not os.path.exists(avatars_folder):
        print(f"❌ 错误: 找不到头像文件夹 {avatars_folder}")
        return []
    
    try:
        with open(json_file, 'r', encoding='utf-8') as file:
            operators_data = json.load(file)
        
        valid_operators = []
        
        for operator in operators_data:
            name = operator.get('姓名', 'Unknown')
            
            if name != 'Unknown':
                # 使用文件名规则查找文件
                filename, file_path = get_avatar_filename(name, avatars_folder)
                
                if filename and file_path:
                    operator['头像本地路径'] = file_path
                    operator['头像文件名'] = filename
                    valid_operators.append(operator)
        
        print(f"✅ 找到 {len(valid_operators)} 个有效的干员头像")
        return valid_operators
        
    except Exception as e:
        print(f"❌ 获取有效干员时出错: {e}")
        return []

def resize_image_for_excel(image_path, target_width=100, target_height=100):
    """
    调整图片尺寸以适合Excel单元格
    """
    try:
        with PILImage.open(image_path) as img:
            # 转换为RGB模式（如果是RGBA）
            if img.mode in ('RGBA', 'LA'):
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            
            # 等比例缩放
            img.thumbnail((target_width, target_height), PILImage.Resampling.LANCZOS)
            
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            img.save(temp_file.name, 'PNG')
            return temp_file.name
    except Exception as e:
        print(f"❌ 处理图片 {image_path} 时出错: {e}")
        return None

def create_single_excel(selected_operators, output_file, version_name, show_details=True):
    """
    创建单个Excel文件
    """
    if show_details:
        print(f"\n📊 创建 {version_name}...")
        print(f"🎯 干员顺序:")
        for i, op in enumerate(selected_operators[:26]):
            print(f"  {i+1:2d}. {op.get('姓名')} - {op.get('头像文件名', '无文件')}")
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"随机干员头像-{version_name}"
    
    # 设置列宽和行高
    cell_width = 15
    cell_height = 75
    
    # 设置前6列的宽度（第一行）
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = cell_width
    
    # 设置前5列的宽度（后续行）
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = cell_width
    
    # 设置行高
    for row in range(1, 6):
        ws.row_dimensions[row].height = cell_height
    
    # 创建边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 创建背景色
    fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    
    # 布局位置定义
    positions = []
    # 第一行：6个位置
    for col in range(1, 7):
        positions.append((1, col))
    # 后续4行：每行5个位置
    for row in range(2, 6):
        for col in range(1, 6):
            positions.append((row, col))
    
    temp_files = []
    successful_inserts = 0
    
    try:
        # 插入头像和信息
        for i, operator in enumerate(selected_operators[:26]):
            if i >= len(positions):
                break
                
            row, col = positions[i]
            cell = ws.cell(row=row, column=col)
            name = operator.get('姓名', 'Unknown')
            
            # 设置单元格基本样式
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 先设置文字（作为备选）
            profession = operator.get('子职业', '')
            cell_text = f"{name}"
            if profession:
                cell_text += f"\n{profession}"
            cell.value = cell_text
            
            # 尝试插入图片
            avatar_path = operator.get('头像本地路径')
            if avatar_path and os.path.exists(avatar_path) and HAS_IMAGE_SUPPORT:
                try:
                    # 调整图片尺寸
                    resized_image_path = resize_image_for_excel(avatar_path, 90, 90)
                    if resized_image_path:
                        temp_files.append(resized_image_path)
                        
                        # 创建Excel图片对象
                        img = Image(resized_image_path)
                        img.width = 90
                        img.height = 90
                        
                        # 设置锚点
                        col_letter = get_column_letter(col)
                        anchor_cell = f"{col_letter}{row}"
                        img.anchor = anchor_cell
                        
                        # 添加图片到工作表
                        ws.add_image(img)
                        
                        # 清空单元格文字（图片已添加）
                        cell.value = ""
                        
                        successful_inserts += 1
                        
                except Exception as e:
                    if show_details:
                        print(f"❌ 插入 {name} 图片时出错: {e}")
    
        if show_details:
            print(f"📊 {version_name} 处理完成:")
            print(f"  - 成功插入图片: {successful_inserts} 个")
            print(f"  - 文字显示: {len(selected_operators[:26]) - successful_inserts} 个")
    
    except Exception as e:
        print(f"❌ 创建 {version_name} 时出错: {e}")
    
    finally:
        # 保存Excel文件
        try:
            wb.save(output_file)
            
            file_size = os.path.getsize(output_file)
            if show_details:
                print(f"✅ {version_name} 已保存: {output_file}")
                print(f"📄 文件大小: {file_size} 字节")
            
            # 清理临时文件
            for temp_file in temp_files:
                try:
                    os.unlink(temp_file)
                except:
                    pass
                    
        except Exception as e:
            print(f"❌ 保存 {version_name} 时出错: {e}")
    
    return successful_inserts

def main():
    """
    主函数
    """
    print("🚀 明日方舟干员头像Excel生成器")
    print("=" * 60)
    
    # 获取资源文件路径
    json_file = get_resource_path('operators_data.json')
    avatars_folder = get_resource_path('avatars')
    
    # 获取exe文件所在目录作为输出目录
    output_dir = check_and_create_output_dir()
    
    # 设置输出文件路径（保存在exe文件同一目录）
    output_file1 = os.path.join(output_dir, '版本A_随机干员头像.xlsx')
    output_file2 = os.path.join(output_dir, '版本B_随机干员头像.xlsx')
    
    print(f"📁 JSON数据文件: {json_file}")
    print(f"📁 头像文件夹: {avatars_folder}")
    print(f"📁 输出目录: {output_dir}")
    print(f"📁 输出文件1: {os.path.basename(output_file1)}")
    print(f"📁 输出文件2: {os.path.basename(output_file2)}")
    
    # 获取所有有效的干员
    valid_operators = get_valid_operators(json_file, avatars_folder)
    
    if len(valid_operators) == 0:
        print("❌ 没有找到有效的头像文件")
        print("请确保以下文件存在:")
        print(f"  - {json_file}")
        print(f"  - {avatars_folder} (包含头像图片)")
        input("按回车键退出...")
        return
    
    # 随机选择26个干员
    if len(valid_operators) >= 26:
        selected_operators = random.sample(valid_operators, 26)
        print(f"🎲 从 {len(valid_operators)} 个有效干员中随机选择了 26 个")
    else:
        selected_operators = valid_operators
        print(f"⚠️ 只有 {len(valid_operators)} 个有效干员，少于26个")
    
    # 显示选中的干员列表
    print(f"\n🎯 选中的26个干员:")
    for i, op in enumerate(selected_operators):
        print(f"  {i+1:2d}. {op.get('姓名')}")
    
    # 创建第一个版本（原始顺序）
    print(f"\n" + "="*50)
    successful_A = create_single_excel(
        selected_operators.copy(), 
        output_file1, 
        "版本A (原始顺序)", 
        show_details=True
    )
    
    # 创建第二个版本（打乱顺序）
    shuffled_operators = selected_operators.copy()
    random.shuffle(shuffled_operators)
    
    print(f"\n" + "="*50)
    successful_B = create_single_excel(
        shuffled_operators, 
        output_file2, 
        "版本B (打乱顺序)", 
        show_details=True
    )
    
    # 显示最终结果
    print(f"\n" + "="*60)
    print(f"🎉 双版本生成完成！")
    print(f"📁 文件1: {output_file1}")
    print(f"📁 文件2: {output_file2}")
    print(f"👥 每个文件包含: {len(selected_operators)} 个干员")
    print(f"🖼️ 版本A 图片数: {successful_A}")
    print(f"🖼️ 版本B 图片数: {successful_B}")
    
    # 显示顺序对比（前10个）
    print(f"\n📋 顺序对比 (前10个):")
    print(f"{'序号':<4} {'版本A':<12} {'版本B':<12}")
    print(f"{'-'*4} {'-'*12} {'-'*12}")
    for i in range(min(10, len(selected_operators))):
        name_a = selected_operators[i].get('姓名', 'Unknown')
        name_b = shuffled_operators[i].get('姓名', 'Unknown')
        print(f"{i+1:<4} {name_a:<12} {name_b:<12}")
    
    if len(selected_operators) > 10:
        print(f"{'...':<4} {'...':<12} {'...':<12}")
    
    print(f"\n🏁 程序执行完成")
    print(f"📂 生成的Excel文件已保存在程序同一目录下")
    input("按回车键退出...")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ 程序运行出错: {e}")
        input("按回车键退出...")

